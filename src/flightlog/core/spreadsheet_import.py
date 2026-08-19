"""
Self-service spreadsheet import (v0.9.8, specs/008-self-service-import).

A pilot uploads their own Excel or CSV file and maps their own column headers to Flightlog
fields via the UI — a generic mapping wizard, not a parser for any specific third-party export
format. Deliberately independent of `core/importer.py`: that module's fixed column positions and
`core/aliases.py` alias tables are this pilot's own legacy-workbook cleanup, not something that
generalises to an arbitrary new pilot's spreadsheet. Reference-data reuse here is exact-string
match only — never fuzzy, same "never guess-match" principle that made the old bulk-IGC matcher
get pulled in v0.8.1.

`run_import()` does one pass for both preview and commit: it always resolves/creates rows inside
a transaction, then either commits or rolls back, so preview and commit can never silently
diverge in what they'd produce (a duplicated ruleset is how that kind of drift usually starts).
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from flightlog.database.models import (
    Flight,
    FlightCategory,
    Glider,
    Harness,
    ImportRun,
    Site,
)

logger = logging.getLogger(__name__)

REQUIRED_FIELDS = ("flight_date", "launch_site")
OPTIONAL_FIELDS = (
    "landing_site",
    "category",
    "glider",
    "harness",
    "duration_min",
    "distance_km",
    "max_alt_m",
    "launch_technique",
    "nickname",
    "notes",
)
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS

# Rows with no mapped category (column unmapped, or blank on that row) land here rather than
# blocking on flights.category_id's NOT NULL constraint — spec.md's Assumptions.
IMPORTED_CATEGORY_NAME = "Imported"

_DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y")


class SpreadsheetError(Exception):
    """The file itself can't be read — wrong type, no header row, corrupt bytes."""


def sha256_hex(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


# ---- reading the raw grid ----


def list_sheet_names(filename: str, data: bytes) -> list[str]:
    """Empty for CSV (no concept of sheets). For Excel, every real worksheet — the mapping UI
    lets the pilot pick one rather than silently assuming the first, since a real workbook (this
    project's own legacy one included) routinely carries several unrelated sheets."""
    if filename.lower().endswith(".csv"):
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        return list(wb.sheetnames)
    except Exception as exc:
        raise SpreadsheetError(f"Could not read this as an Excel file: {exc}") from exc


def _read_grid(filename: str, data: bytes, sheet: str | None) -> list[list[object]]:
    """Row 0 is the header row. Every other row is raw cell values (str for CSV; whatever
    openpyxl returns — str/float/date/datetime/None — for Excel)."""
    if filename.lower().endswith(".csv"):
        return _read_csv_grid(data)
    return _read_xlsx_grid(data, sheet)


def _read_csv_grid(data: bytes) -> list[list[object]]:
    text: str | None = None
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise SpreadsheetError("Could not decode the file as UTF-8 or Windows-1252 text")

    try:
        dialect = csv.Sniffer().sniff(text[:4096])
    except csv.Error:
        dialect = csv.excel  # comma-delimited fallback
    rows = list(csv.reader(io.StringIO(text), dialect))
    return [[cell if cell != "" else None for cell in row] for row in rows]


def _read_xlsx_grid(data: bytes, sheet: str | None) -> list[list[object]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise SpreadsheetError(f"Sheet not found: {sheet}")
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    except SpreadsheetError:
        raise
    except Exception as exc:
        raise SpreadsheetError(f"Could not read this as an Excel file: {exc}") from exc


def _resolved_sheet_key(filename: str, data: bytes, sheet: str | None) -> str:
    """The concrete worksheet an import_key must be scoped to, so two different sheets of the
    same file (identical sha256) never collide on `upload:<sha256>:<row>` — resolves an implicit
    "first sheet" selection to that sheet's real name, so the key is stable regardless of
    whether the caller passed `sheet` explicitly or left it to default. Empty for CSV, which has
    no concept of sheets."""
    if filename.lower().endswith(".csv"):
        return ""
    if sheet is not None:
        return sheet
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    return wb.worksheets[0].title


def _dedupe_headers(raw: list[object]) -> list[str]:
    """Blank cells become distinct `Column N` placeholders; genuine duplicate names (two
    columns literally both called "Notes") get a ` (2)`, ` (3)`, ... suffix so the mapping UI
    never shows two identical, indistinguishable choices that silently collapse to whichever
    one a dict comprehension happened to keep last."""
    seen: dict[str, int] = {}
    headers: list[str] = []
    for i, cell in enumerate(raw):
        name = str(cell).strip() if cell not in (None, "") else f"Column {i + 1}"
        seen[name] = seen.get(name, 0) + 1
        headers.append(name if seen[name] == 1 else f"{name} ({seen[name]})")
    return headers


def read_columns(filename: str, data: bytes, sheet: str | None = None) -> list[dict]:
    """Header names + up to 3 sample values per column, for the mapping UI."""
    grid = _read_grid(filename, data, sheet)
    if not grid or all(c is None for c in grid[0]):
        raise SpreadsheetError("No header row found")

    headers = _dedupe_headers(grid[0])
    samples: list[list[str]] = [[] for _ in headers]
    for row in grid[1:6]:
        for i in range(len(headers)):
            if i < len(row) and row[i] is not None and len(samples[i]) < 3:
                samples[i].append(str(row[i]))
    return [{"name": h, "samples": samples[i]} for i, h in enumerate(headers)]


# ---- per-row parsing (pure — no DB) ----


@dataclass
class ParsedRow:
    row_index: int  # 1-based, header excluded — "row 5" in a user-facing message
    ok: bool
    error: str | None = None
    flight_date: date | None = None
    launch_site: str | None = None
    landing_site: str | None = None
    category: str | None = None
    glider: str | None = None
    harness: str | None = None
    duration_min: int | None = None
    distance_km: float | None = None
    max_alt_m: int | None = None
    launch_technique: str | None = None
    nickname: str | None = None
    notes: str | None = None


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_int(value: object) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return round(float(value))
    except (TypeError, ValueError):
        return None


def _to_float(value: object) -> float | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_launch_technique(value: object) -> str | None:
    text = _to_str(value)
    if text is None:
        return None
    lowered = text.lower()
    if lowered in ("forward", "f"):
        return "forward"
    if lowered in ("reverse", "r"):
        return "reverse"
    return None


def parse_rows(
    filename: str, data: bytes, mapping: dict[str, str], sheet: str | None = None
) -> list[ParsedRow]:
    """`mapping` is {flightlog_field: source_column_header}. Every row is classified — never
    silently dropped (FR-005/FR-012)."""
    grid = _read_grid(filename, data, sheet)
    if not grid:
        return []
    headers = _dedupe_headers(grid[0])
    header_index = {h: i for i, h in enumerate(headers)}

    missing = [f for f in mapping if mapping[f] not in header_index]
    if missing:
        raise SpreadsheetError(f"Mapped column(s) not found in file: {', '.join(missing)}")

    def cell(row: list[object], field_name: str) -> object:
        column = mapping.get(field_name)
        if column is None:
            return None
        idx = header_index[column]
        return row[idx] if idx < len(row) else None

    parsed: list[ParsedRow] = []
    for i, row in enumerate(grid[1:], start=1):
        if all(c is None for c in row):
            continue  # a genuinely blank row is not a row at all, never reported as an error

        flight_date = _to_date(cell(row, "flight_date"))
        launch_site = _to_str(cell(row, "launch_site"))

        if flight_date is None:
            parsed.append(ParsedRow(row_index=i, ok=False, error="Missing or unreadable date"))
            continue
        if launch_site is None:
            parsed.append(ParsedRow(row_index=i, ok=False, error="Missing launch site"))
            continue

        parsed.append(
            ParsedRow(
                row_index=i,
                ok=True,
                flight_date=flight_date,
                launch_site=launch_site,
                landing_site=_to_str(cell(row, "landing_site")),
                category=_to_str(cell(row, "category")),
                glider=_to_str(cell(row, "glider")),
                harness=_to_str(cell(row, "harness")),
                duration_min=_to_int(cell(row, "duration_min")),
                distance_km=_to_float(cell(row, "distance_km")),
                max_alt_m=_to_int(cell(row, "max_alt_m")),
                launch_technique=_to_launch_technique(cell(row, "launch_technique")),
                nickname=_to_str(cell(row, "nickname")),
                notes=_to_str(cell(row, "notes")),
            )
        )
    return parsed


# ---- reference-data resolution (exact match only — never fuzzy) ----


def _get_or_create_site(
    db: Session,
    cache: dict[str, Site],
    owner_id: str,
    name: str,
    *,
    is_launch: bool,
    is_landing: bool,
    import_run_id: str,
    created: list[Site],
) -> Site:
    if name in cache:
        site = cache[name]
        if is_launch and not site.is_launch:
            site.is_launch = True
        if is_landing and not site.is_landing:
            site.is_landing = True
        return site

    existing = db.execute(
        select(Site).where(Site.owner_id == owner_id, Site.name == name)
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing
        if is_launch and not existing.is_launch:
            existing.is_launch = True
        if is_landing and not existing.is_landing:
            existing.is_landing = True
        return existing

    site = Site(
        owner_id=owner_id,
        name=name,
        is_launch=is_launch,
        is_landing=is_landing,
        import_run_id=import_run_id,
    )
    db.add(site)
    db.flush()
    cache[name] = site
    created.append(site)
    return site


def _get_or_create_glider(
    db: Session,
    cache: dict[str, Glider],
    owner_id: str,
    name: str,
    import_run_id: str,
    created: list[Glider],
) -> Glider:
    if name in cache:
        return cache[name]
    existing = db.execute(
        select(Glider).where(Glider.owner_id == owner_id, Glider.model == name)
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing
        return existing
    glider = Glider(owner_id=owner_id, brand="", model=name, import_run_id=import_run_id)
    db.add(glider)
    db.flush()
    cache[name] = glider
    created.append(glider)
    return glider


def _get_or_create_harness(
    db: Session,
    cache: dict[str, Harness],
    owner_id: str,
    name: str,
    import_run_id: str,
    created: list[Harness],
) -> Harness:
    if name in cache:
        return cache[name]
    existing = db.execute(
        select(Harness).where(Harness.owner_id == owner_id, Harness.model == name)
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing
        return existing
    harness = Harness(owner_id=owner_id, brand="", model=name, import_run_id=import_run_id)
    db.add(harness)
    db.flush()
    cache[name] = harness
    created.append(harness)
    return harness


def _get_or_create_category(
    db: Session,
    cache: dict[str, FlightCategory],
    owner_id: str,
    name: str,
    import_run_id: str,
    created: list[FlightCategory],
) -> FlightCategory:
    if name in cache:
        return cache[name]
    existing = db.execute(
        select(FlightCategory).where(
            FlightCategory.owner_id == owner_id, FlightCategory.name == name
        )
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing
        return existing
    slug = name.lower().replace("&", "-and-").replace(" ", "-")
    existing_slug = db.execute(
        select(FlightCategory).where(
            FlightCategory.owner_id == owner_id, FlightCategory.slug == slug
        )
    ).scalar_one_or_none()
    if existing_slug is not None:
        cache[name] = existing_slug
        return existing_slug
    category = FlightCategory(
        owner_id=owner_id,
        name=name,
        slug=slug,
        import_run_id=import_run_id,
    )
    db.add(category)
    db.flush()
    cache[name] = category
    created.append(category)
    return category


# ---- orchestration ----


@dataclass
class ImportOutcome:
    row_count: int = 0
    imported_count: int = 0
    already_imported_count: int = 0
    skipped_count: int = 0
    errors: list[dict] = field(default_factory=list)  # [{row, reason}]
    new_sites: list[str] = field(default_factory=list)
    new_gliders: list[str] = field(default_factory=list)
    new_harnesses: list[str] = field(default_factory=list)
    new_categories: list[str] = field(default_factory=list)
    import_run_id: str | None = None  # only set when committed


def run_import(
    db: Session,
    owner_id: str,
    filename: str,
    data: bytes,
    mapping: dict[str, str],
    *,
    commit: bool,
    sheet: str | None = None,
) -> ImportOutcome:
    """One pass, used for both preview (`commit=False`, rolled back) and the real thing
    (`commit=True`) — so the two paths can never silently diverge in what they'd produce."""
    missing_required = [f for f in REQUIRED_FIELDS if f not in mapping]
    if missing_required:
        raise SpreadsheetError(f"Required field(s) not mapped: {', '.join(missing_required)}")

    rows = parse_rows(filename, data, mapping, sheet)
    sha256 = sha256_hex(data)
    sheet_key = _resolved_sheet_key(filename, data, sheet)
    outcome = ImportOutcome(row_count=len(rows))

    run = ImportRun(
        owner_id=owner_id,
        source_filename=filename,
        column_mapping=_encode_mapping(mapping),
        row_count=len(rows),
        imported_count=0,
        skipped_count=0,
    )
    db.add(run)
    db.flush()

    site_cache: dict[str, Site] = {}
    glider_cache: dict[str, Glider] = {}
    harness_cache: dict[str, Harness] = {}
    category_cache: dict[str, FlightCategory] = {}
    new_sites: list[Site] = []
    new_gliders: list[Glider] = []
    new_harnesses: list[Harness] = []
    new_categories: list[FlightCategory] = []

    for row in rows:
        if not row.ok:
            outcome.skipped_count += 1
            outcome.errors.append({"row": row.row_index, "reason": row.error})
            continue

        import_key = f"upload:{sha256}:{sheet_key}:{row.row_index}"
        existing_flight = db.execute(
            select(Flight.id).where(Flight.owner_id == owner_id, Flight.import_key == import_key)
        ).scalar_one_or_none()
        if existing_flight is not None:
            outcome.already_imported_count += 1
            continue

        launch_site = _get_or_create_site(
            db,
            site_cache,
            owner_id,
            row.launch_site,
            is_launch=True,
            is_landing=False,
            import_run_id=run.id,
            created=new_sites,
        )
        landing_site = None
        if row.landing_site:
            landing_site = _get_or_create_site(
                db,
                site_cache,
                owner_id,
                row.landing_site,
                is_launch=False,
                is_landing=True,
                import_run_id=run.id,
                created=new_sites,
            )
        category = _get_or_create_category(
            db,
            category_cache,
            owner_id,
            row.category or IMPORTED_CATEGORY_NAME,
            run.id,
            new_categories,
        )
        glider = None
        if row.glider:
            glider = _get_or_create_glider(
                db, glider_cache, owner_id, row.glider, run.id, new_gliders
            )
        harness = None
        if row.harness:
            harness = _get_or_create_harness(
                db, harness_cache, owner_id, row.harness, run.id, new_harnesses
            )

        db.add(
            Flight(
                owner_id=owner_id,
                flight_date=row.flight_date,
                launch_site_id=launch_site.id,
                landing_site_id=landing_site.id if landing_site else None,
                category_id=category.id,
                glider_id=glider.id if glider else None,
                harness_id=harness.id if harness else None,
                duration_min=row.duration_min,
                distance_km=row.distance_km,
                max_alt_m=row.max_alt_m,
                launch_technique=row.launch_technique,
                nickname=row.nickname,
                notes=row.notes,
                import_key=import_key,
                import_run_id=run.id,
            )
        )
        outcome.imported_count += 1

    run.imported_count = outcome.imported_count
    run.skipped_count = outcome.skipped_count
    outcome.new_sites = [s.name for s in new_sites]
    outcome.new_gliders = [g.model for g in new_gliders]
    outcome.new_harnesses = [h.model for h in new_harnesses]
    outcome.new_categories = [c.name for c in new_categories]

    if commit:
        db.commit()
        outcome.import_run_id = run.id
        logger.info(
            "Import run %s: %d/%d rows imported, %d skipped, %d already-imported (owner=%s)",
            run.id,
            outcome.imported_count,
            outcome.row_count,
            outcome.skipped_count,
            outcome.already_imported_count,
            owner_id,
        )
    else:
        db.rollback()

    return outcome


def _encode_mapping(mapping: dict[str, str]) -> str:
    return json.dumps(mapping, sort_keys=True)


def decode_mapping(column_mapping: str) -> dict[str, str]:
    return json.loads(column_mapping)


# ---- undo ----


@dataclass
class UndoOutcome:
    flights_deleted: int = 0
    flights_kept: int = 0
    sites_deleted: int = 0
    gliders_deleted: int = 0
    harnesses_deleted: int = 0
    categories_deleted: int = 0
    reference_rows_kept: int = 0


def undo_import(db: Session, run: ImportRun) -> UndoOutcome:
    """Deletes everything this run created that is still untouched and unreferenced elsewhere;
    anything that isn't safe to delete is kept, just untagged. Deletes `run` itself last."""
    outcome = UndoOutcome()
    owner_id = run.owner_id

    flights = db.execute(select(Flight).where(Flight.import_run_id == run.id)).scalars().all()
    for flight in flights:
        if flight.updated_at is None:
            db.delete(flight)
            outcome.flights_deleted += 1
        else:
            flight.import_run_id = None
            outcome.flights_kept += 1
    db.flush()

    def _sweep(model, fk_columns, counter_name: str) -> None:
        rows = db.execute(select(model).where(model.import_run_id == run.id)).scalars().all()
        for row in rows:
            still_used = db.execute(
                select(Flight.id)
                .where(Flight.owner_id == owner_id, or_(*[fk == row.id for fk in fk_columns]))
                .limit(1)
            ).scalar_one_or_none()
            if still_used is None:
                db.delete(row)
                setattr(outcome, counter_name, getattr(outcome, counter_name) + 1)
            else:
                row.import_run_id = None
                outcome.reference_rows_kept += 1

    # A site can be referenced as either a launch or a landing site — both columns must be
    # checked before deciding a tagged site is unused (checking only one would wrongly delete
    # a site that's still in use through the other role).
    _sweep(Site, [Flight.launch_site_id, Flight.landing_site_id], "sites_deleted")
    _sweep(Glider, [Flight.glider_id], "gliders_deleted")
    _sweep(Harness, [Flight.harness_id], "harnesses_deleted")
    _sweep(FlightCategory, [Flight.category_id], "categories_deleted")

    db.flush()
    db.delete(run)
    db.commit()
    logger.info(
        "Import run undone: %s (flights deleted=%d kept=%d, reference rows kept=%d)",
        run.id,
        outcome.flights_deleted,
        outcome.flights_kept,
        outcome.reference_rows_kept,
    )
    return outcome
