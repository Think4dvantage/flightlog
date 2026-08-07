"""
One-shot Excel importer.

    python -m flightlog.core.importer              # dry-run (default) against the legacy workbook
    python -m flightlog.core.importer --write       # commits
    python -m flightlog.core.importer --path FILE   # a different workbook (e.g. for testing)

Reads the `Flugbuch` sheet row by row and creates regions, sites, gliders, harnesses,
flight_categories and flights for the single existing pilot account (spec.md's Assumptions:
"There is exactly one pilot account in the system today"). Every reference is resolved through
`core/aliases.py`; nothing unresolved is guessed.

A row that can't resolve a launch site or a category is skipped and reported — never written with
a placeholder, since both columns are NOT NULL on `Flight` (data-model.md). An unresolved landing
site, glider or harness does not block the flight — it is written with that field left null, and
the mismatch is reported (FR-014's "never silently dropped" means never dropped *without a trace*,
not that the whole flight is blocked on optional gear).

Idempotent: flights are looked up by `(owner_id, import_key)`, everything else by
`(owner_id, canonical_name)` — a second run against the same file changes nothing (FR-012).

Buddy names recognized in `Kommentar` text (against `core/aliases.py`'s `KNOWN_BUDDY_NAMES`) are
surfaced in the report as proposals only — no `buddies` row is ever created automatically (FR-017).
"""

from __future__ import annotations

import argparse
import logging
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from flightlog.core.aliases import (
    CANONICAL_CATEGORIES,
    CANONICAL_GLIDERS,
    CANONICAL_HARNESSES,
    CANONICAL_LANDINGS,
    CANONICAL_LAUNCHES,
    CATEGORY_ALIASES,
    CATEGORY_FLAGS,
    GLIDER_ALIASES,
    HARNESS_ALIASES,
    KNOWN_BUDDY_NAMES,
    LAUNCH_TYPE_MAP,
    SITE_ALIASES,
    SITE_REGION,
)
from flightlog.database.models import Flight, FlightCategory, Glider, Harness, Region, Site, User

logger = logging.getLogger(__name__)

FLUGBUCH_SHEET = "Flugbuch"

# 0-based column indices, verbatim from the real workbook's header row (research.md).
COL_DATE = 0
COL_LAUNCH = 1
COL_LAUNCH_ELEV = 2
COL_LANDING = 3
COL_LANDING_ELEV = 4
COL_HOEHE_DIFF = 5
COL_FLUGZEIT = 6
COL_DISTANZ = 7
COL_KATEGORIE = 8
COL_MAX_ALT = 9
COL_ALTGAIN = 10
COL_SCHIRM = 11
COL_GURTZEUG = 12
COL_STARTART = 13
COL_KOMMENTAR = 14


@dataclass
class ImportReport:
    rows_read: int = 0
    flights_written: int = 0
    flights_skipped_existing: int = 0
    flights_skipped_unresolved: list[dict] = field(default_factory=list)
    regions_written: int = 0
    sites_written: int = 0
    gliders_written: int = 0
    harnesses_written: int = 0
    categories_written: int = 0
    # kind -> [(source_value, canonical_value), ...]
    alias_hits: dict[str, list[tuple[str, str]]] = field(
        default_factory=lambda: {"site": [], "glider": [], "harness": [], "category": []}
    )
    # kind -> {value: count} — matched nothing, canonical or alias
    unresolved: dict[str, Counter] = field(
        default_factory=lambda: {
            "launch_site": Counter(),
            "landing_site": Counter(),
            "category": Counter(),
            "glider": Counter(),
            "harness": Counter(),
        }
    )
    # region_name -> flight count, recomputed from the rows this run actually resolved
    region_counts: Counter = field(default_factory=Counter)
    # region_name -> {"computed": N, "sheet": M} where they disagree — FR-015
    region_mismatches: dict[str, dict] = field(default_factory=dict)
    # [{row, computed_alt_gain_m, sheet_altgain, delta}] — FR-016
    altgain_mismatches: list[dict] = field(default_factory=list)
    # name -> flight count mentioning them — proposals only, never auto-created (FR-017)
    buddy_proposals: Counter = field(default_factory=Counter)


def _resolve(
    value: str | None,
    canonical: list[str],
    aliases: dict[str, str],
    kind: str,
    report: ImportReport,
) -> str | None:
    """
    Returns the canonical name, or None if value is blank or matches nothing.

    Records an alias hit on `report` as a side effect when a variant spelling resolves.
    Does NOT record unresolved misses — callers do that themselves, since the same `kind`
    (e.g. "site") covers multiple report buckets (launch_site vs landing_site) that need
    different context attached to the miss.
    """
    if not value:
        return None
    if value in canonical:
        return value
    if value in aliases:
        canonical_value = aliases[value]
        report.alias_hits[kind].append((value, canonical_value))
        return canonical_value
    return None


def _get_or_create_region(
    db: Session, cache: dict[str, str], name: str | None, report: ImportReport
) -> str | None:
    if name is None:
        return None
    if name in cache:
        return cache[name]
    existing = db.execute(select(Region.id).where(Region.name == name)).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing
        return existing
    region = Region(name=name, sort_order=len(cache))
    db.add(region)
    db.flush()
    cache[name] = region.id
    report.regions_written += 1
    return region.id


def _get_or_create_site(
    db: Session,
    cache: dict[str, str],
    owner_id: str,
    name: str,
    is_launch: bool,
    is_landing: bool,
    elevation_m: int | None,
    region_id: str | None,
    report: ImportReport,
) -> str:
    if name in cache:
        site_id = cache[name]
        # A site already seen as a landing that now also appears as a launch (or vice
        # versa) gets both flags set — same real place used two ways (architecture.md).
        site = db.get(Site, site_id)
        if is_launch and not site.is_launch:
            site.is_launch = True
        if is_landing and not site.is_landing:
            site.is_landing = True
        return site_id

    existing = db.execute(
        select(Site).where(Site.owner_id == owner_id, Site.name == name)
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing.id
        if is_launch and not existing.is_launch:
            existing.is_launch = True
        if is_landing and not existing.is_landing:
            existing.is_landing = True
        return existing.id

    site = Site(
        owner_id=owner_id,
        name=name,
        is_launch=is_launch,
        is_landing=is_landing,
        elevation_m=elevation_m,
        region_id=region_id,
        coord_source=None,
    )
    db.add(site)
    db.flush()
    cache[name] = site.id
    report.sites_written += 1
    return site.id


def _get_or_create_glider(
    db: Session, cache: dict[str, str], owner_id: str, name: str, report: ImportReport
) -> str:
    # Legacy names carry brand+model as one string (e.g. "(Ragnar) Advance Epsilon 9 28");
    # splitting them is a v0.3+ data-quality task, not this import's job. Store verbatim
    # in `model`, leave `brand` blank rather than inventing a split that isn't there yet.
    if name in cache:
        return cache[name]
    existing = db.execute(
        select(Glider).where(Glider.owner_id == owner_id, Glider.model == name)
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing.id
        return existing.id

    glider = Glider(owner_id=owner_id, brand="", model=name)
    db.add(glider)
    db.flush()
    cache[name] = glider.id
    report.gliders_written += 1
    return glider.id


def _get_or_create_harness(
    db: Session, cache: dict[str, str], owner_id: str, name: str, report: ImportReport
) -> str:
    if name in cache:
        return cache[name]
    existing = db.execute(
        select(Harness).where(Harness.owner_id == owner_id, Harness.model == name)
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing.id
        return existing.id

    harness = Harness(owner_id=owner_id, brand="", model=name)
    db.add(harness)
    db.flush()
    cache[name] = harness.id
    report.harnesses_written += 1
    return harness.id


def _get_or_create_category(
    db: Session, cache: dict[str, str], owner_id: str, name: str, report: ImportReport
) -> str:
    if name in cache:
        return cache[name]
    existing = db.execute(
        select(FlightCategory).where(
            FlightCategory.owner_id == owner_id, FlightCategory.name == name
        )
    ).scalar_one_or_none()
    if existing is not None:
        cache[name] = existing.id
        return existing.id

    flags = CATEGORY_FLAGS.get(name, {"is_hike_fly": False, "is_training": False})
    category = FlightCategory(
        owner_id=owner_id,
        name=name,
        slug=name.lower().replace("&", "-and-").replace(" ", "-"),
        sort_order=len(cache),
        **flags,
    )
    db.add(category)
    db.flush()
    cache[name] = category.id
    report.categories_written += 1
    return category.id


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_int(value) -> int | None:
    if value is None:
        return None
    try:
        return round(float(value))
    except (TypeError, ValueError):
        return None


_BUDDY_NAME_PATTERNS = {
    name: re.compile(rf"\b{re.escape(name)}\b", re.IGNORECASE) for name in KNOWN_BUDDY_NAMES
}


def _find_buddy_mentions(comment: str | None) -> list[str]:
    """Case-insensitive, word-boundary matching against KNOWN_BUDDY_NAMES — never creates
    anything, just names who is proposed (FR-017)."""
    if not comment:
        return []
    return [name for name, pattern in _BUDDY_NAME_PATTERNS.items() if pattern.search(comment)]


def _read_sheet_region_counts(wb) -> dict[str, int]:
    """
    Reads the "Flight Area" block from Übersicht's Total column (col C) — the workbook's OWN
    region totals, read as data, never recomputed from formulas. Used only as the comparison
    baseline for FR-015; this project's own region counts always come from SITE_REGION.

    Übersicht is not present in every fixture workbook (region reconciliation is optional
    per-file) — returns {} if the sheet is missing rather than raising.
    """
    if "Übersicht" not in wb.sheetnames:
        return {}
    ws = wb["Übersicht"]

    counts: dict[str, int] = {}
    in_flight_area = False
    for row in ws.iter_rows(values_only=True):
        name, _b, total = row[0], row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None
        if name is None and isinstance(total, str) and total == "Flight Area":
            in_flight_area = True
            continue
        if name is None and isinstance(total, str):
            # Any other section marker ends the Flight Area block.
            if in_flight_area:
                break
            continue
        if in_flight_area and isinstance(name, str) and isinstance(total, (int, float)):
            counts[name] = int(total)
    return counts


def run_import(db: Session, path: str, owner_id: str, write: bool = False) -> ImportReport:
    """Reads `Flugbuch` from `path`, resolves every reference, and either previews (default)
    or commits (`write=True`) the resulting sites/gliders/harnesses/categories/flights."""
    report = ImportReport()

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[FLUGBUCH_SHEET]

    region_cache: dict[str, str] = {}
    site_cache: dict[str, str] = {}
    glider_cache: dict[str, str] = {}
    harness_cache: dict[str, str] = {}
    category_cache: dict[str, str] = {}

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[COL_DATE] is None:
            continue
        report.rows_read += 1
        import_key = f"xlsx:{excel_row_number}"

        existing_flight = db.execute(
            select(Flight.id).where(Flight.owner_id == owner_id, Flight.import_key == import_key)
        ).scalar_one_or_none()
        if existing_flight is not None:
            report.flights_skipped_existing += 1
            continue

        launch_name = _resolve(row[COL_LAUNCH], CANONICAL_LAUNCHES, SITE_ALIASES, "site", report)
        if launch_name is None:
            report.unresolved["launch_site"][row[COL_LAUNCH]] += 1
            report.flights_skipped_unresolved.append(
                {"row": excel_row_number, "reason": "launch_site", "value": row[COL_LAUNCH]}
            )
            continue

        category_name = _resolve(
            row[COL_KATEGORIE], CANONICAL_CATEGORIES, CATEGORY_ALIASES, "category", report
        )
        if category_name is None:
            report.unresolved["category"][row[COL_KATEGORIE]] += 1
            report.flights_skipped_unresolved.append(
                {"row": excel_row_number, "reason": "category", "value": row[COL_KATEGORIE]}
            )
            continue

        landing_name = _resolve(row[COL_LANDING], CANONICAL_LANDINGS, SITE_ALIASES, "site", report)
        if row[COL_LANDING] and landing_name is None:
            report.unresolved["landing_site"][row[COL_LANDING]] += 1

        glider_name = _resolve(row[COL_SCHIRM], CANONICAL_GLIDERS, GLIDER_ALIASES, "glider", report)
        if row[COL_SCHIRM] and glider_name is None:
            report.unresolved["glider"][row[COL_SCHIRM]] += 1

        harness_name = _resolve(
            row[COL_GURTZEUG], CANONICAL_HARNESSES, HARNESS_ALIASES, "harness", report
        )
        if row[COL_GURTZEUG] and harness_name is None:
            report.unresolved["harness"][row[COL_GURTZEUG]] += 1

        # Region and altitude-figure checks run whether or not this pass writes — a
        # trustworthy dry-run report (FR-015/FR-016) must not require --write first.
        region_name = SITE_REGION.get(launch_name)
        if region_name is not None:
            report.region_counts[region_name] += 1

        launch_elev = _to_int(row[COL_LAUNCH_ELEV])
        max_alt = _to_int(row[COL_MAX_ALT])
        sheet_altgain = _to_int(row[COL_ALTGAIN])
        if launch_elev is not None and max_alt is not None and sheet_altgain is not None:
            computed_alt_gain = max_alt - launch_elev
            if computed_alt_gain != sheet_altgain:
                report.altgain_mismatches.append(
                    {
                        "row": excel_row_number,
                        "computed_alt_gain_m": computed_alt_gain,
                        "sheet_altgain": sheet_altgain,
                        "delta": computed_alt_gain - sheet_altgain,
                    }
                )

        for name in _find_buddy_mentions(row[COL_KOMMENTAR]):
            report.buddy_proposals[name] += 1

        if not write:
            continue

        region_id = _get_or_create_region(db, region_cache, SITE_REGION.get(launch_name), report)
        launch_site_id = _get_or_create_site(
            db,
            site_cache,
            owner_id,
            launch_name,
            True,
            False,
            _to_int(row[COL_LAUNCH_ELEV]),
            region_id,
            report,
        )
        landing_site_id = None
        if landing_name is not None:
            landing_site_id = _get_or_create_site(
                db,
                site_cache,
                owner_id,
                landing_name,
                False,
                True,
                _to_int(row[COL_LANDING_ELEV]),
                None,
                report,
            )
        category_id = _get_or_create_category(db, category_cache, owner_id, category_name, report)
        glider_id = (
            _get_or_create_glider(db, glider_cache, owner_id, glider_name, report)
            if glider_name
            else None
        )
        harness_id = (
            _get_or_create_harness(db, harness_cache, owner_id, harness_name, report)
            if harness_name
            else None
        )

        startart_raw = (row[COL_STARTART] or "").strip().lower()
        launch_technique = LAUNCH_TYPE_MAP.get(startart_raw)

        flight = Flight(
            owner_id=owner_id,
            flight_date=_to_date(row[COL_DATE]),
            launch_site_id=launch_site_id,
            landing_site_id=landing_site_id,
            category_id=category_id,
            glider_id=glider_id,
            harness_id=harness_id,
            duration_min=_to_int(row[COL_FLUGZEIT]),
            distance_km=row[COL_DISTANZ],
            max_alt_m=_to_int(row[COL_MAX_ALT]),
            launch_technique=launch_technique,
            notes=row[COL_KOMMENTAR],
            import_key=import_key,
        )
        db.add(flight)
        report.flights_written += 1

    sheet_region_counts = _read_sheet_region_counts(wb)
    all_region_names = set(report.region_counts) | set(sheet_region_counts)
    for region_name in all_region_names:
        computed = report.region_counts.get(region_name, 0)
        sheet_value = sheet_region_counts.get(region_name)
        if sheet_value is not None and computed != sheet_value:
            report.region_mismatches[region_name] = {"computed": computed, "sheet": sheet_value}

    if write:
        db.commit()

    return report


def _get_single_owner(db: Session) -> User:
    users = db.execute(select(User)).scalars().all()
    if len(users) != 1:
        raise RuntimeError(
            f"Expected exactly one pilot account, found {len(users)} — the importer needs "
            "spec.md's stated assumption to hold. Aborting."
        )
    return users[0]


def _print_report(report: ImportReport, write: bool) -> None:
    mode = "WRITE" if write else "DRY-RUN"
    print(f"--- Import report ({mode}) ---")
    print(f"Rows read: {report.rows_read}")
    print(f"Flights written: {report.flights_written}")
    print(f"Flights skipped (already imported): {report.flights_skipped_existing}")
    print(f"Flights skipped (unresolved): {len(report.flights_skipped_unresolved)}")
    for entry in report.flights_skipped_unresolved:
        print(f"  row {entry['row']}: unresolved {entry['reason']} = {entry['value']!r}")
    print(f"Regions written: {report.regions_written}")
    print(f"Sites written: {report.sites_written}")
    print(f"Gliders written: {report.gliders_written}")
    print(f"Harnesses written: {report.harnesses_written}")
    print(f"Categories written: {report.categories_written}")
    for kind, hits in report.alias_hits.items():
        for source, canonical in hits:
            print(f"  alias: {kind} {source!r} -> {canonical!r}")
    for kind, counter in report.unresolved.items():
        for value, count in counter.items():
            print(f"  unresolved {kind}: {value!r} x{count}")
    for region_name, counts in report.region_mismatches.items():
        print(
            f"  region mismatch: {region_name} computed={counts['computed']} "
            f"sheet={counts['sheet']}"
        )
    for entry in report.altgain_mismatches:
        print(
            f"  altgain mismatch: row {entry['row']} computed={entry['computed_alt_gain_m']} "
            f"sheet={entry['sheet_altgain']} delta={entry['delta']}"
        )
    for name, count in report.buddy_proposals.items():
        print(f"  buddy proposal: {name} mentioned in {count} flight(s)")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", default="olddata/Flugbuch.xlsx")
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args(argv)

    from flightlog.config import load_config
    from flightlog.database.db import init_db

    cfg = load_config()
    engine = init_db(cfg.database.path)
    from sqlalchemy.orm import Session

    with Session(engine) as db:
        owner = _get_single_owner(db)
        logger.info("Importing %s for owner=%s (write=%s)", args.path, owner.id, args.write)
        report = run_import(db, args.path, owner.id, write=args.write)
        _print_report(report, args.write)


if __name__ == "__main__":
    main()
