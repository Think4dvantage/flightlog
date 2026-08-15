"""
One-shot import of the legacy workbook's four remaining sheets: `Fitnessprogramm` (hikes),
`Groundhandling`, `Tandemflüge` (tandem flights taken as a passenger), and `Ziele` (goals).

    python -m flightlog.core.secondary_import              # dry-run (default)
    python -m flightlog.core.secondary_import --write       # commits
    python -m flightlog.core.secondary_import --path FILE   # a different workbook

Idempotent exactly like `core/importer.py`: each row gets an `import_key` of `"<sheet>:<row>"`,
looked up by `(owner_id, import_key)` before writing anything — a second run against the same
file changes nothing.

Hikes/ground-handling/tandem-flights are import-and-view only (no API write path exists for
them). Goals are the one type that stays editable afterward — this importer's `import_key`
guard only ever creates the initial rows; every subsequent read/write on a goal goes through
the normal `/api/goals` CRUD router, never back through this module.
"""

from __future__ import annotations

import argparse
import logging
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from flightlog.database.models import (
    Flight,
    FlightCategory,
    Goal,
    GroundhandlingSession,
    Hike,
    TandemFlight,
)

logger = logging.getLogger(__name__)

FITNESSPROGRAMM_SHEET = "Fitnessprogramm"
GROUNDHANDLING_SHEET = "Groundhandling"
TANDEMFLUEGE_SHEET = "Tandemflüge"
ZIELE_SHEET = "Ziele"

# 0-based column indices, verbatim from each sheet's real header row (research.md).
FP_COL_DATUM = 0
FP_COL_START = 1
FP_COL_ZIEL = 2
FP_COL_STEIGUNG = 3
FP_COL_GEFAELLE = 4
FP_COL_DISTANZ = 5
FP_COL_ZEIT = 6
FP_COL_ROUTE = 7
FP_COL_AIRTIME = 8
FP_COL_LANDEPLATZ = 9

GH_COL_DATUM = 0
GH_COL_ORT = 1
GH_COL_DAUER = 2
GH_COL_KOMMENTAR = 3

TF_COL_DATUM = 0
TF_COL_START = 1
TF_COL_LANDUNG = 2
TF_COL_PILOT = 3
TF_COL_KOMMENTAR = 4
TF_COL_KOSTEN = 5

# Ziele reports ~505 columns wide per row, but every column past the 8th is a leftover Excel
# formatting artifact — None on every real row (research.md). Read only these 8 by position.
ZI_COL_TITEL = 0
ZI_COL_WETTERLAGE = 1
ZI_COL_LEVEL = 2
ZI_COL_KATEGORIE = 3
ZI_COL_BESCHREIBUNG = 4
ZI_COL_LINKS = 5
ZI_COL_SAISON = 6
ZI_COL_STATUS = 7


@dataclass
class SecondaryImportReport:
    hikes_read: int = 0
    hikes_written: int = 0
    hikes_skipped_existing: int = 0
    hikes_linked: int = 0
    hikes_ambiguous: list[dict] = field(default_factory=list)
    groundhandling_read: int = 0
    groundhandling_written: int = 0
    groundhandling_skipped_existing: int = 0
    tandem_flights_read: int = 0
    tandem_flights_written: int = 0
    tandem_flights_skipped_existing: int = 0
    goals_read: int = 0
    goals_written: int = 0
    goals_skipped_existing: int = 0


def _to_date(value):
    from datetime import date, datetime

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_int(value):
    if value is None:
        return None
    try:
        return round(float(value))
    except (TypeError, ValueError):
        return None


def _find_hike_flight_link(
    db: Session, owner_id: str, hike_date, report: SecondaryImportReport, excel_row_number: int
) -> str | None:
    """Link only when exactly one Hike&Fly-category flight shares this hike's date — never
    guessed (research.md's ambiguity rule, matching the IGC bulk-match precedent)."""
    candidates = (
        db.execute(
            select(Flight.id)
            .join(FlightCategory, Flight.category_id == FlightCategory.id)
            .where(
                Flight.owner_id == owner_id,
                Flight.flight_date == hike_date,
                FlightCategory.is_hike_fly.is_(True),
            )
        )
        .scalars()
        .all()
    )
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        report.hikes_ambiguous.append(
            {"row": excel_row_number, "date": str(hike_date), "candidate_count": len(candidates)}
        )
    return None


def _import_hikes(
    db: Session, wb, owner_id: str, write: bool, report: SecondaryImportReport
) -> None:
    if FITNESSPROGRAMM_SHEET not in wb.sheetnames:
        return
    ws = wb[FITNESSPROGRAMM_SHEET]

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[FP_COL_DATUM] is None:
            continue
        report.hikes_read += 1
        import_key = f"fitnessprogramm:{excel_row_number}"

        existing = db.execute(
            select(Hike.id).where(Hike.owner_id == owner_id, Hike.import_key == import_key)
        ).scalar_one_or_none()
        if existing is not None:
            report.hikes_skipped_existing += 1
            continue

        if not write:
            continue

        hike_date = _to_date(row[FP_COL_DATUM])
        became_a_flight = row[FP_COL_AIRTIME] is not None or row[FP_COL_LANDEPLATZ] is not None
        flight_id = (
            _find_hike_flight_link(db, owner_id, hike_date, report, excel_row_number)
            if became_a_flight
            else None
        )
        if flight_id is not None:
            report.hikes_linked += 1

        hike = Hike(
            owner_id=owner_id,
            import_key=import_key,
            hike_date=hike_date,
            start_place=row[FP_COL_START],
            destination_place=row[FP_COL_ZIEL],
            ascent_m=_to_int(row[FP_COL_STEIGUNG]),
            descent_m=_to_int(row[FP_COL_GEFAELLE]),
            distance_km=row[FP_COL_DISTANZ],
            duration_min=_to_int(row[FP_COL_ZEIT]),
            route_description=row[FP_COL_ROUTE],
            flight_id=flight_id,
        )
        db.add(hike)
        report.hikes_written += 1


def _import_groundhandling(
    db: Session, wb, owner_id: str, write: bool, report: SecondaryImportReport
) -> None:
    if GROUNDHANDLING_SHEET not in wb.sheetnames:
        return
    ws = wb[GROUNDHANDLING_SHEET]

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[GH_COL_DATUM] is None:
            continue
        report.groundhandling_read += 1
        import_key = f"groundhandling:{excel_row_number}"

        existing = db.execute(
            select(GroundhandlingSession.id).where(
                GroundhandlingSession.owner_id == owner_id,
                GroundhandlingSession.import_key == import_key,
            )
        ).scalar_one_or_none()
        if existing is not None:
            report.groundhandling_skipped_existing += 1
            continue

        if not write:
            continue

        db.add(
            GroundhandlingSession(
                owner_id=owner_id,
                import_key=import_key,
                session_date=_to_date(row[GH_COL_DATUM]),
                place=row[GH_COL_ORT],
                duration_min=_to_int(row[GH_COL_DAUER]),
                comment=row[GH_COL_KOMMENTAR],
            )
        )
        report.groundhandling_written += 1


def _import_tandem_flights(
    db: Session, wb, owner_id: str, write: bool, report: SecondaryImportReport
) -> None:
    if TANDEMFLUEGE_SHEET not in wb.sheetnames:
        return
    ws = wb[TANDEMFLUEGE_SHEET]

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[TF_COL_DATUM] is None:
            continue
        report.tandem_flights_read += 1
        import_key = f"tandemfluege:{excel_row_number}"

        existing = db.execute(
            select(TandemFlight.id).where(
                TandemFlight.owner_id == owner_id, TandemFlight.import_key == import_key
            )
        ).scalar_one_or_none()
        if existing is not None:
            report.tandem_flights_skipped_existing += 1
            continue

        if not write:
            continue

        db.add(
            TandemFlight(
                owner_id=owner_id,
                import_key=import_key,
                flight_date=_to_date(row[TF_COL_DATUM]),
                launch_place=row[TF_COL_START],
                landing_place=row[TF_COL_LANDUNG],
                tandem_operator=row[TF_COL_PILOT],
                comment=row[TF_COL_KOMMENTAR],
                cost=row[TF_COL_KOSTEN],
            )
        )
        report.tandem_flights_written += 1


def _import_goals(
    db: Session, wb, owner_id: str, write: bool, report: SecondaryImportReport
) -> None:
    if ZIELE_SHEET not in wb.sheetnames:
        return
    ws = wb[ZIELE_SHEET]

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[ZI_COL_TITEL] is None:
            continue
        report.goals_read += 1
        import_key = f"ziele:{excel_row_number}"

        existing = db.execute(
            select(Goal.id).where(Goal.owner_id == owner_id, Goal.import_key == import_key)
        ).scalar_one_or_none()
        if existing is not None:
            report.goals_skipped_existing += 1
            continue

        if not write:
            continue

        status = row[ZI_COL_STATUS] or "open"
        db.add(
            Goal(
                owner_id=owner_id,
                import_key=import_key,
                title=row[ZI_COL_TITEL],
                wind_direction=row[ZI_COL_WETTERLAGE],
                difficulty=row[ZI_COL_LEVEL],
                category=row[ZI_COL_KATEGORIE],
                description=row[ZI_COL_BESCHREIBUNG],
                links=row[ZI_COL_LINKS],
                target_season=str(row[ZI_COL_SAISON]) if row[ZI_COL_SAISON] is not None else None,
                status=status,
            )
        )
        report.goals_written += 1


def run_secondary_import(
    db: Session, path: str, owner_id: str, write: bool = False
) -> SecondaryImportReport:
    """Reads Fitnessprogramm/Groundhandling/Tandemflüge/Ziele from `path` and either previews
    (default) or commits (`write=True`) hikes/groundhandling_sessions/tandem_flights/goals."""
    report = SecondaryImportReport()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    _import_hikes(db, wb, owner_id, write, report)
    _import_groundhandling(db, wb, owner_id, write, report)
    _import_tandem_flights(db, wb, owner_id, write, report)
    _import_goals(db, wb, owner_id, write, report)

    if write:
        db.commit()

    return report


def _get_single_owner(db: Session):
    from flightlog.database.models import User

    users = db.execute(select(User)).scalars().all()
    if len(users) != 1:
        raise RuntimeError(
            f"Expected exactly one pilot account, found {len(users)} — the importer needs "
            "spec.md's stated assumption to hold. Aborting."
        )
    return users[0]


def _print_report(report: SecondaryImportReport, write: bool) -> None:
    mode = "WRITE" if write else "DRY-RUN"
    print(f"--- Secondary sheets import report ({mode}) ---")
    print(f"Hikes read: {report.hikes_read}")
    print(f"Hikes written: {report.hikes_written}")
    print(f"Hikes skipped (already imported): {report.hikes_skipped_existing}")
    print(f"Hikes linked to a flight: {report.hikes_linked}")
    for entry in report.hikes_ambiguous:
        print(
            f"  ambiguous hike link: row {entry['row']} date={entry['date']} "
            f"candidates={entry['candidate_count']}"
        )
    print(f"Groundhandling sessions read: {report.groundhandling_read}")
    print(f"Groundhandling sessions written: {report.groundhandling_written}")
    print(
        "Groundhandling sessions skipped (already imported): "
        f"{report.groundhandling_skipped_existing}"
    )
    print(f"Tandem flights read: {report.tandem_flights_read}")
    print(f"Tandem flights written: {report.tandem_flights_written}")
    print(f"Tandem flights skipped (already imported): {report.tandem_flights_skipped_existing}")
    print(f"Goals read: {report.goals_read}")
    print(f"Goals written: {report.goals_written}")
    print(f"Goals skipped (already imported): {report.goals_skipped_existing}")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", default="olddata/Flugbuch.xlsx")
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args(argv)

    from sqlalchemy.orm import Session

    from flightlog.config import load_config
    from flightlog.database.db import init_db

    cfg = load_config()
    engine = init_db(cfg.database.path)

    with Session(engine) as db:
        owner = _get_single_owner(db)
        logger.info(
            "Importing secondary sheets from %s for owner=%s (write=%s)",
            args.path,
            owner.id,
            args.write,
        )
        report = run_secondary_import(db, args.path, owner.id, write=args.write)
        _print_report(report, args.write)


if __name__ == "__main__":
    main()
