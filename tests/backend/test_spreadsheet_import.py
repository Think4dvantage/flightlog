"""Self-service spreadsheet import (v0.9.8, specs/008-self-service-import): upload/map, preview
vs. commit, exact-match reference-data reuse, idempotent re-upload, and undo."""

from __future__ import annotations

import io
import json

import openpyxl
import pytest
from sqlalchemy import select

from flightlog.core.spreadsheet_import import _to_date
from flightlog.database.models import Flight, FlightCategory, Site, User

CSV_HEADER = "Date,Launch,Landing,Category,Duration\n"


def _csv(rows: str) -> bytes:
    return (CSV_HEADER + rows).encode("utf-8")


GOOD_CSV = _csv(
    "2024-08-15,Amisbühl,Interlaken,Thermal,45\n2024-08-16,Amisbühl,Interlaken,Thermal,30\n"
)

MAPPING = json.dumps(
    {
        "flight_date": "Date",
        "launch_site": "Launch",
        "landing_site": "Landing",
        "category": "Category",
        "duration_min": "Duration",
    }
)


def _xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Launch", "Landing", "Category", "Duration"])
    ws.append(["2024-09-01", "Beatenberg", "Interlaken", "Soaring", "60"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _multi_sheet_xlsx() -> bytes:
    """A summary sheet first, the real flight data on a second sheet — mirrors this project's
    own legacy workbook (Flugbuch is one of six sheets, never sheet 1 by assumption)."""
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Total flights", "Total hours"])
    summary.append([2, 1.25])

    flights = wb.create_sheet("Flights")
    flights.append(["Date", "Launch", "Landing", "Category", "Duration"])
    flights.append(["2024-10-01", "Niederhorn", "Interlaken", "Thermal", "40"])
    flights.append(["2024-10-02", "Niederhorn", "Interlaken", "Thermal", "35"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


DUPLICATE_HEADER_CSV = (
    "Date,Launch,Notes,Notes\n2024-08-15,Amisbühl,first note,second note\n"
).encode()


@pytest.fixture
def headers(make_token):
    return make_token()


def test_to_date_handles_common_formats():
    assert _to_date("2024-08-15").isoformat() == "2024-08-15"
    assert _to_date("15.08.2024").isoformat() == "2024-08-15"
    assert _to_date("08/15/2024").isoformat() == "2024-08-15"
    assert _to_date("not a date") is None
    assert _to_date(None) is None


async def test_columns_lists_headers_and_samples(client, headers):
    resp = await client.post(
        "/api/imports/columns",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        headers=headers,
    )
    assert resp.status_code == 200
    names = [c["name"] for c in resp.json()["columns"]]
    assert names == ["Date", "Launch", "Landing", "Category", "Duration"]
    date_col = resp.json()["columns"][0]
    assert date_col["samples"] == ["2024-08-15", "2024-08-16"]


async def test_preview_does_not_write_anything(client, headers, db_session):
    resp = await client.post(
        "/api/imports/preview",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["row_count"] == 2
    assert body["imported_count"] == 2
    assert body["new_sites"] == ["Amisbühl", "Interlaken"]
    assert body["new_categories"] == ["Thermal"]

    db_session.expire_all()
    assert db_session.execute(select(Flight)).scalars().all() == []


async def test_commit_creates_flights_and_reference_rows(client, headers, db_session):
    resp = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["imported_count"] == 2
    assert body["import_run_id"]

    db_session.expire_all()
    flights = db_session.execute(select(Flight).order_by(Flight.flight_date)).scalars().all()
    sites = db_session.execute(select(Site)).scalars().all()
    categories = db_session.execute(select(FlightCategory)).scalars().all()

    assert len(flights) == 2
    assert {s.name for s in sites} == {"Amisbühl", "Interlaken"}
    assert {c.name for c in categories} == {"Thermal"}
    assert flights[0].duration_min == 45
    assert flights[0].import_run_id == body["import_run_id"]


async def test_xlsx_upload_also_works(client, headers):
    resp = await client.post(
        "/api/imports/commit",
        files={
            "file": (
                "log.xlsx",
                _xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert resp.status_code == 201
    assert resp.json()["imported_count"] == 1


async def test_multi_sheet_xlsx_lists_sheets_and_defaults_to_the_first(client, headers):
    resp = await client.post(
        "/api/imports/columns",
        files={"file": ("log.xlsx", _multi_sheet_xlsx(), "application/octet-stream")},
        headers=headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["sheet_names"] == ["Summary", "Flights"]
    # Defaults to the first sheet — the summary sheet's own headers, not the flight data.
    assert [c["name"] for c in body["columns"]] == ["Total flights", "Total hours"]


async def test_committing_a_non_first_sheet_uses_that_sheets_data(client, headers, db_session):
    mapping = json.dumps(
        {
            "flight_date": "Date",
            "launch_site": "Launch",
            "landing_site": "Landing",
            "category": "Category",
            "duration_min": "Duration",
        }
    )
    resp = await client.post(
        "/api/imports/commit",
        files={"file": ("log.xlsx", _multi_sheet_xlsx(), "application/octet-stream")},
        data={"mapping": mapping, "sheet": "Flights"},
        headers=headers,
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["imported_count"] == 2
    assert body["new_sites"] == ["Niederhorn", "Interlaken"]

    flights = db_session.execute(select(Flight)).scalars().all()
    assert len(flights) == 2


async def test_duplicate_column_headers_are_disambiguated_not_collapsed(client, headers):
    resp = await client.post(
        "/api/imports/columns",
        files={"file": ("log.csv", DUPLICATE_HEADER_CSV, "text/csv")},
        headers=headers,
    )
    assert resp.status_code == 200
    names = [c["name"] for c in resp.json()["columns"]]
    assert names == ["Date", "Launch", "Notes", "Notes (2)"]

    mapping = json.dumps({"flight_date": "Date", "launch_site": "Launch", "notes": "Notes (2)"})
    commit = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", DUPLICATE_HEADER_CSV, "text/csv")},
        data={"mapping": mapping},
        headers=headers,
    )
    assert commit.status_code == 201
    assert commit.json()["imported_count"] == 1


async def test_same_file_different_sheet_is_not_treated_as_a_duplicate(client, headers, db_session):
    """Two sheets of one file share a sha256 — the idempotency key must still tell them apart."""
    file_bytes = _multi_sheet_xlsx()
    mapping = json.dumps(
        {
            "flight_date": "Date",
            "launch_site": "Launch",
            "landing_site": "Landing",
            "category": "Category",
            "duration_min": "Duration",
        }
    )
    # "Flights" sheet only has 2 rows; add a matching second sheet with the same shape so
    # this test exercises row_index collisions across sheets, not just sheet content.
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    other = wb.create_sheet("MoreFlights")
    other.append(["Date", "Launch", "Landing", "Category", "Duration"])
    other.append(["2024-11-01", "Niederhorn", "Interlaken", "Thermal", "50"])
    other.append(["2024-11-02", "Niederhorn", "Interlaken", "Thermal", "55"])
    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    first = await client.post(
        "/api/imports/commit",
        files={"file": ("log.xlsx", file_bytes, "application/octet-stream")},
        data={"mapping": mapping, "sheet": "Flights"},
        headers=headers,
    )
    assert first.json()["imported_count"] == 2

    second = await client.post(
        "/api/imports/commit",
        files={"file": ("log.xlsx", file_bytes, "application/octet-stream")},
        data={"mapping": mapping, "sheet": "MoreFlights"},
        headers=headers,
    )
    assert second.status_code == 201
    body = second.json()
    assert body["already_imported_count"] == 0
    assert body["imported_count"] == 2

    assert len(db_session.execute(select(Flight)).scalars().all()) == 4


async def test_unparseable_row_is_reported_not_dropped(client, headers):
    csv_bytes = _csv("not-a-date,Amisbühl,,,\n2024-08-16,,,,\n2024-08-17,Amisbühl,,Thermal,20\n")
    resp = await client.post(
        "/api/imports/preview",
        files={"file": ("log.csv", csv_bytes, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["row_count"] == 3
    assert body["imported_count"] == 1
    assert body["skipped_count"] == 2
    reasons = {e["row"]: e["reason"] for e in body["errors"]}
    assert reasons[1] == "Missing or unreadable date"
    assert reasons[2] == "Missing launch site"


async def test_unmapped_category_falls_back_to_imported_bucket(client, headers, db_session):
    mapping = json.dumps({"flight_date": "Date", "launch_site": "Launch"})
    resp = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": mapping},
        headers=headers,
    )
    assert resp.status_code == 201

    db_session.expire_all()
    categories = db_session.execute(select(FlightCategory)).scalars().all()
    assert [c.name for c in categories] == ["Imported"]
    assert len(db_session.execute(select(Flight)).scalars().all()) == 2


async def test_reuses_exact_matching_existing_reference_data(client, headers, db_session):
    user = db_session.execute(select(User)).scalar_one()
    existing = Site(owner_id=user.id, name="Amisbühl", is_launch=True)
    db_session.add(existing)
    db_session.commit()
    existing_id = existing.id

    resp = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert resp.status_code == 201
    assert resp.json()["new_sites"] == ["Interlaken"]  # Amisbühl reused, not recreated

    db_session.expire_all()
    sites = db_session.execute(select(Site).where(Site.name == "Amisbühl")).scalars().all()
    assert len(sites) == 1
    assert sites[0].id == existing_id
    assert sites[0].is_launch  # unchanged — GOOD_CSV never uses Amisbühl as a landing site


async def test_reuploading_the_same_file_is_idempotent(client, headers, db_session):
    first = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert first.json()["imported_count"] == 2

    second = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    assert second.status_code == 201
    body = second.json()
    assert body["imported_count"] == 0
    assert body["already_imported_count"] == 2

    db_session.expire_all()
    assert len(db_session.execute(select(Flight)).scalars().all()) == 2


async def test_required_field_not_mapped_is_422(client, headers):
    mapping = json.dumps({"flight_date": "Date"})  # launch_site missing
    resp = await client.post(
        "/api/imports/preview",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": mapping},
        headers=headers,
    )
    assert resp.status_code == 422
    assert resp.json()["error"]["code"] == "VALIDATION_FAILED"


async def test_non_spreadsheet_extension_rejected(client, headers):
    resp = await client.post(
        "/api/imports/columns",
        files={"file": ("log.txt", b"whatever", "text/plain")},
        headers=headers,
    )
    assert resp.status_code == 422


async def test_undo_removes_created_flights_and_reference_rows(client, headers, db_session):
    commit = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    run_id = commit.json()["import_run_id"]

    undo = await client.delete(f"/api/imports/{run_id}", headers=headers)
    assert undo.status_code == 200
    body = undo.json()
    assert body["flights_deleted"] == 2
    assert body["sites_deleted"] == 2
    assert body["categories_deleted"] == 1

    db_session.expire_all()
    assert db_session.execute(select(Flight)).scalars().all() == []
    assert db_session.execute(select(Site)).scalars().all() == []
    assert db_session.execute(select(FlightCategory)).scalars().all() == []


async def test_undo_keeps_an_edited_flight_and_still_used_reference_rows(
    client, headers, db_session
):
    commit = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    run_id = commit.json()["import_run_id"]

    db_session.expire_all()
    flights = db_session.execute(select(Flight).order_by(Flight.flight_date)).scalars().all()
    edited_flight_id = flights[0].id
    kept_launch_site_id = flights[0].launch_site_id

    edit = await client.put(
        f"/api/flights/{edited_flight_id}",
        json={"nickname": "edited after import"},
        headers=headers,
    )
    assert edit.status_code == 200

    undo = await client.delete(f"/api/imports/{run_id}", headers=headers)
    assert undo.status_code == 200
    body = undo.json()
    assert body["flights_deleted"] == 1
    assert body["flights_kept"] == 1

    db_session.expire_all()
    remaining = db_session.execute(select(Flight)).scalars().all()
    assert len(remaining) == 1
    assert remaining[0].id == edited_flight_id
    assert remaining[0].import_run_id is None
    # The site the kept flight still launches from must survive undo even though it was
    # tagged by the same run.
    assert db_session.get(Site, kept_launch_site_id) is not None


async def test_another_users_import_run_is_404_not_403(client, headers, make_token):
    commit = await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    run_id = commit.json()["import_run_id"]

    other_headers = make_token(email="other@example.com")
    resp = await client.delete(f"/api/imports/{run_id}", headers=other_headers)
    assert resp.status_code == 404


async def test_list_import_runs(client, headers):
    await client.post(
        "/api/imports/commit",
        files={"file": ("log.csv", GOOD_CSV, "text/csv")},
        data={"mapping": MAPPING},
        headers=headers,
    )
    resp = await client.get("/api/imports", headers=headers)
    assert resp.status_code == 200
    runs = resp.json()
    assert len(runs) == 1
    assert runs[0]["source_filename"] == "log.csv"
    assert runs[0]["imported_count"] == 2
